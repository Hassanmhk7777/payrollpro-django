"""
paie/excel_cnss.py - Générateur d'export CNSS format officiel Maroc
Intégration avec le système PayrollPro existant
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from decimal import Decimal
from datetime import datetime
from django.db.models import Sum, Count
from django.utils import timezone


class ExporteurCNSS:
    """
    Générateur d'export CNSS au format Bordereau de Déclaration des Salaires (BDS)
    Compatible avec la législation marocaine 2025
    """
    
    def __init__(self):
        # Informations entreprise (à configurer via admin Django)
        self.info_entreprise = {
            'raison_sociale': 'PayrollPro SARL',
            'n_affiliation_cnss': '2521470',
            'n_patente': '12345678', 
            'n_rc': '987654',
            'adresse': '123 Boulevard Zerktouni',
            'ville': 'Casablanca',
            'code_postal': '20000'
        }
        
        # Taux officiels CNSS 2025
        self.taux = {
            'cnss_salarie': Decimal('4.48'),
            'amo_salarie': Decimal('2.26'),
            'cnss_patronal': Decimal('20.48'),  # NOUVEAU
            'amo_patronal': Decimal('1.85'),    # NOUVEAU
            'formation_prof': Decimal('1.60'),  # NOUVEAU
            'plafond_cnss': Decimal('6000.00')
        }
    
    def generer_bds_mensuel(self, mois, annee):
        """
        Génère le BDS mensuel pour tous les employés avec bulletins calculés
        
        Args:
            mois (int): Mois (1-12)
            annee (int): Année (ex: 2025)
            
        Returns:
            Workbook: Fichier Excel prêt pour export
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"BDS_{mois:02d}_{annee}"
        
        # 1. Créer l'en-tête entreprise
        self._creer_entete_entreprise(ws, mois, annee)
        
        # 2. Récupérer les bulletins du mois
        from .models import BulletinPaie
        bulletins = BulletinPaie.objects.filter(
            mois=mois,
            annee=annee
        ).select_related('employe').order_by('employe__matricule')
        
        if not bulletins.exists():
            self._ajouter_message_aucun_bulletin(ws)
            return wb
        
        # 3. Créer l'en-tête des colonnes
        self._creer_entete_colonnes(ws, ligne_debut=8)
        
        # 4. Remplir les données des employés
        ligne_actuelle = 9
        totaux = self._initialiser_totaux()
        
        for bulletin in bulletins:
            ligne_actuelle = self._ajouter_ligne_employe(
                ws, bulletin, ligne_actuelle, totaux
            )
        
        # 5. Ajouter les totaux
        self._ajouter_totaux(ws, totaux, ligne_actuelle + 1)
        
        # 6. Ajouter les calculs patronaux
        self._ajouter_calculs_patronaux(ws, totaux, ligne_actuelle + 5)
        
        # 7. Formater le document
        self._formater_document(ws)
        
        return wb
    
    def _creer_entete_entreprise(self, ws, mois, annee):
        """Crée l'en-tête avec les informations de l'entreprise"""
        
        # Titre principal
        ws.merge_cells('A1:Q1')
        ws['A1'] = "BORDEREAU DE DÉCLARATION DES SALAIRES (BDS)"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Période
        ws.merge_cells('A2:Q2')
        mois_noms = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                     'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
        ws['A2'] = f"PÉRIODE : {mois_noms[mois]} {annee}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Informations entreprise
        ligne = 4
        infos = [
            f"RAISON SOCIALE : {self.info_entreprise['raison_sociale']}",
            f"N° AFFILIATION CNSS : {self.info_entreprise['n_affiliation_cnss']}",
            f"N° PATENTE : {self.info_entreprise['n_patente']}",
            f"ADRESSE : {self.info_entreprise['adresse']}, {self.info_entreprise['ville']}"
        ]
        
        for info in infos:
            ws[f'A{ligne}'] = info
            ws[f'A{ligne}'].font = Font(bold=True)
            ligne += 1
    
    def _creer_entete_colonnes(self, ws, ligne_debut):
        """Crée l'en-tête des colonnes du tableau"""
        
        colonnes = [
            ('A', 'MATRICULE'),
            ('B', 'NOM & PRÉNOM'),
            ('C', 'FONCTION'),
            ('D', 'TYPE CONTRAT'),
            ('E', 'DATE EMBAUCHE'),
            ('F', 'SIT. FAMILIALE'),
            ('G', 'NB ENFANTS'),
            ('H', 'JOURS TRAVAILLÉS'),
            ('I', 'SALAIRE BRUT GLOBAL'),
            ('J', 'SALAIRE BRUT IMPOSABLE'),
            ('K', 'SALAIRE BRUT PLAFONNÉ'),
            ('L', 'CNSS (4.48%)'),
            ('M', 'AMO (2.26%)'),
            ('N', 'CIMR (6%)'),
            ('O', 'IMPÔT REVENU'),
            ('P', 'AVANCES'),
            ('Q', 'NET À PAYER')
        ]
        
        for col, titre in colonnes:
            cellule = f'{col}{ligne_debut}'
            ws[cellule] = titre
            ws[cellule].font = Font(bold=True, color='FFFFFF')
            ws[cellule].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            ws[cellule].alignment = Alignment(horizontal='center', wrap_text=True)
            ws[cellule].border = self._get_border()
    
    def _ajouter_ligne_employe(self, ws, bulletin, ligne, totaux):
        """Ajoute une ligne pour un employé dans le BDS"""
        
        employe = bulletin.employe
        
        # Données de base
        ws[f'A{ligne}'] = employe.matricule
        ws[f'B{ligne}'] = employe.nom_complet()
        ws[f'C{ligne}'] = employe.fonction
        ws[f'D{ligne}'] = 'CDI'  # À améliorer si vous avez cette info
        ws[f'E{ligne}'] = employe.date_embauche.strftime('%d/%m/%Y')
        ws[f'F{ligne}'] = employe.get_situation_familiale_display()
        ws[f'G{ligne}'] = employe.nombre_enfants
        ws[f'H{ligne}'] = 26  # Jours ouvrables standard
        
        # Montants financiers
        ws[f'I{ligne}'] = float(bulletin.salaire_brut_imposable + bulletin.total_primes)
        ws[f'J{ligne}'] = float(bulletin.salaire_brut_imposable)
        
        # Salaire brut plafonné pour CNSS
        brut_plafonne = min(bulletin.salaire_brut_imposable, self.taux['plafond_cnss'])
        ws[f'K{ligne}'] = float(brut_plafonne)
        
        # Cotisations
        ws[f'L{ligne}'] = float(bulletin.cotisation_cnss)
        ws[f'M{ligne}'] = float(bulletin.cotisation_amo)
        ws[f'N{ligne}'] = float(bulletin.cotisation_cimr)
        ws[f'O{ligne}'] = float(bulletin.impot_revenu)
        ws[f'P{ligne}'] = float(bulletin.total_avances)
        ws[f'Q{ligne}'] = float(bulletin.net_a_payer)
        
        # Mettre à jour les totaux
        self._mettre_a_jour_totaux(totaux, bulletin, brut_plafonne)
        
        # Formater la ligne
        for col in 'ABCDEFGHIJKLMNOPQ':
            ws[f'{col}{ligne}'].border = self._get_border()
            if col in 'IJKLMNOPQ':  # Colonnes numériques
                ws[f'{col}{ligne}'].number_format = '#,##0.00'
                ws[f'{col}{ligne}'].alignment = Alignment(horizontal='right')
        
        return ligne + 1
    
    def _ajouter_totaux(self, ws, totaux, ligne):
        """Ajoute les lignes de totaux"""
        
        # Ligne de totaux salariaux
        ws[f'A{ligne}'] = "TOTAUX"
        ws[f'A{ligne}'].font = Font(bold=True)
        
        colonnes_totaux = [
            ('I', totaux['salaire_brut_global']),
            ('J', totaux['salaire_brut_imposable']),
            ('K', totaux['salaire_brut_plafonne']),
            ('L', totaux['cotisation_cnss']),
            ('M', totaux['cotisation_amo']),
            ('N', totaux['cotisation_cimr']),
            ('O', totaux['impot_revenu']),
            ('P', totaux['avances']),
            ('Q', totaux['net_a_payer'])
        ]
        
        for col, montant in colonnes_totaux:
            ws[f'{col}{ligne}'] = float(montant)
            ws[f'{col}{ligne}'].font = Font(bold=True)
            ws[f'{col}{ligne}'].fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
            ws[f'{col}{ligne}'].number_format = '#,##0.00'
            ws[f'{col}{ligne}'].border = self._get_border()
    
    def _ajouter_calculs_patronaux(self, ws, totaux, ligne):
        """Ajoute le tableau des charges patronales"""
        
        # Titre
        ws[f'A{ligne}'] = "CHARGES PATRONALES"
        ws[f'A{ligne}'].font = Font(bold=True, size=12)
        ligne += 1
        
        # En-têtes
        ws[f'A{ligne}'] = "NATURE"
        ws[f'B{ligne}'] = "BASE"
        ws[f'C{ligne}'] = "TAUX"
        ws[f'D{ligne}'] = "MONTANT"
        
        for col in 'ABCD':
            ws[f'{col}{ligne}'].font = Font(bold=True)
            ws[f'{col}{ligne}'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        ligne += 1
        
        # Calculs patronaux
        base_cnss = totaux['salaire_brut_plafonne']
        base_amo = totaux['salaire_brut_imposable']
        
        charges = [
            ('CNSS Patronal', base_cnss, self.taux['cnss_patronal'], 
             base_cnss * self.taux['cnss_patronal'] / 100),
            ('AMO Patronal', base_amo, self.taux['amo_patronal'],
             base_amo * self.taux['amo_patronal'] / 100),
            ('Formation Prof.', base_cnss, self.taux['formation_prof'],
             base_cnss * self.taux['formation_prof'] / 100)
        ]
        
        total_charges = Decimal('0')
        for nature, base, taux, montant in charges:
            ws[f'A{ligne}'] = nature
            ws[f'B{ligne}'] = float(base)
            ws[f'C{ligne}'] = f"{taux}%"
            ws[f'D{ligne}'] = float(montant)
            
            total_charges += montant
            ligne += 1
        
        # Total charges
        ws[f'A{ligne}'] = "TOTAL CHARGES PATRONALES"
        ws[f'D{ligne}'] = float(total_charges)
        ws[f'A{ligne}'].font = Font(bold=True)
        ws[f'D{ligne}'].font = Font(bold=True)
        ws[f'D{ligne}'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    def _initialiser_totaux(self):
        """Initialise la structure des totaux"""
        return {
            'salaire_brut_global': Decimal('0'),
            'salaire_brut_imposable': Decimal('0'), 
            'salaire_brut_plafonne': Decimal('0'),
            'cotisation_cnss': Decimal('0'),
            'cotisation_amo': Decimal('0'),
            'cotisation_cimr': Decimal('0'),
            'impot_revenu': Decimal('0'),
            'avances': Decimal('0'),
            'net_a_payer': Decimal('0')
        }
    
    def _mettre_a_jour_totaux(self, totaux, bulletin, brut_plafonne):
        """Met à jour les totaux avec les données d'un bulletin"""
        totaux['salaire_brut_global'] += bulletin.salaire_brut_imposable + bulletin.total_primes
        totaux['salaire_brut_imposable'] += bulletin.salaire_brut_imposable
        totaux['salaire_brut_plafonne'] += brut_plafonne
        totaux['cotisation_cnss'] += bulletin.cotisation_cnss
        totaux['cotisation_amo'] += bulletin.cotisation_amo
        totaux['cotisation_cimr'] += bulletin.cotisation_cimr
        totaux['impot_revenu'] += bulletin.impot_revenu
        totaux['avances'] += bulletin.total_avances
        totaux['net_a_payer'] += bulletin.net_a_payer
    
    def _ajouter_message_aucun_bulletin(self, ws):
        """Ajoute un message si aucun bulletin n'est trouvé"""
        ws['A10'] = "⚠️ AUCUN BULLETIN DE PAIE CALCULÉ POUR CETTE PÉRIODE"
        ws['A10'].font = Font(bold=True, color='FF0000', size=14)
        ws['A11'] = "Veuillez d'abord calculer la paie mensuelle avant de générer l'export CNSS."
    
    def _formater_document(self, ws):
        """Applique le formatage global au document"""
        
        # Ajuster la largeur des colonnes
        largeurs = {
            'A': 12, 'B': 20, 'C': 15, 'D': 12, 'E': 12,
            'F': 10, 'G': 8, 'H': 8, 'I': 12, 'J': 12,
            'K': 12, 'L': 10, 'M': 10, 'N': 10, 'O': 10,
            'P': 10, 'Q': 12
        }
        
        for col, largeur in largeurs.items():
            ws.column_dimensions[col].width = largeur
        
        # Hauteur des lignes d'en-tête
        ws.row_dimensions[8].height = 30
    
    def _get_border(self):
        """Retourne un style de bordure standard"""
        thin_border = Side(border_style="thin", color="000000")
        return Border(left=thin_border, right=thin_border, 
                     top=thin_border, bottom=thin_border)
    
    def exporter_vers_reponse_http(self, mois, annee):
        """
        Génère le fichier Excel et retourne une réponse HTTP Django
        
        Args:
            mois (int): Mois
            annee (int): Année
            
        Returns:
            HttpResponse: Réponse avec fichier Excel
        """
        from django.http import HttpResponse
        from io import BytesIO
        
        # Générer le fichier
        wb = self.generer_bds_mensuel(mois, annee)
        
        # Créer la réponse HTTP
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"BDS_CNSS_{mois:02d}_{annee}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response