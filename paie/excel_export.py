# paie/excel_export.py
"""
Système d'export Excel avancé pour PayrollPro
Basé sur les meilleures pratiques 2025 avec openpyxl
"""
import io
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Sum, Count, Avg

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExporteurExcelPayrollPro:
    """
    Générateur d'exports Excel pour le système de paie
    Compatible avec votre structure existante
    """
    
    def __init__(self):
        self.company_name = "PayrollPro - Système de Paie"
        
        # Styles réutilisables (performance optimisée)
        self.styles = {}
        
    def _init_styles(self, workbook):
        """Initialise les styles une fois pour éviter les problèmes de mémoire"""
        if not self.styles:
            self.styles = {
                'title': Font(name='Calibri', size=16, bold=True, color='1F4E79'),
                'header': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
                'header_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
                'currency': '#,##0.00" DH"',
                'date_format': 'DD/MM/YYYY',
                'center': Alignment(horizontal='center'),
                'border_thin': Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            }
    
    def export_bulletin_individuel(self, bulletin):
        """
        Export d'un bulletin de paie individuel au format Excel
        Compatible avec votre modèle BulletinPaie existant
        """
        if not OPENPYXL_AVAILABLE:
            return None
            
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Bulletin_{bulletin.employe.matricule}"
        
        self._init_styles(workbook)
        
        # En-tête société
        worksheet['A1'] = self.company_name
        worksheet['A1'].font = self.styles['title']
        worksheet.merge_cells('A1:F1')
        
        worksheet['A2'] = f"BULLETIN DE PAIE - {bulletin.periode_formatted()}"
        worksheet['A2'].font = Font(name='Calibri', size=14, bold=True)
        worksheet.merge_cells('A2:F2')
        
        # Informations employé (ligne 4-8)
        row = 4
        employe_data = [
            ('Nom et Prénom:', bulletin.employe.nom_complet()),
            ('Matricule:', bulletin.employe.matricule),
            ('Fonction:', bulletin.employe.fonction),
            ('Date embauche:', bulletin.employe.date_embauche.strftime('%d/%m/%Y')),
            ('Situation familiale:', bulletin.employe.get_situation_familiale_display())
        ]
        
        for label, value in employe_data:
            worksheet[f'A{row}'] = label
            worksheet[f'A{row}'].font = Font(bold=True)
            worksheet[f'B{row}'] = value
            row += 1
        
        # Section Absences (si présentes)
        row += 1
        from .calculs import CalculateurPaie
        calculateur = CalculateurPaie()
        absences_info = calculateur.calculer_absences(
            bulletin.employe, bulletin.mois, bulletin.annee
        )
        
        if absences_info['jours_deduits'] > 0:
            worksheet[f'A{row}'] = "ABSENCES ET DÉDUCTIONS"
            worksheet[f'A{row}'].font = Font(bold=True, color='D32F2F')
            row += 1
            
            # En-têtes absences
            headers_absences = ['Type', 'Du', 'Au', 'Jours', 'Motif']
            for col, header in enumerate(headers_absences, 1):
                cell = worksheet.cell(row=row, column=col, value=header)
                cell.font = self.styles['header']
                cell.fill = self.styles['header_fill']
            row += 1
            
            # Détails des absences
            for detail in absences_info['absences_details']:
                worksheet[f'A{row}'] = detail['type']
                worksheet[f'B{row}'] = detail['debut'].strftime('%d/%m/%Y')
                worksheet[f'C{row}'] = detail['fin'].strftime('%d/%m/%Y')
                worksheet[f'D{row}'] = detail['jours']
                worksheet[f'E{row}'] = detail['motif'] or '—'
                row += 1
            
            # Total déduction
            worksheet[f'A{row}'] = "TOTAL DÉDUCTION"
            worksheet[f'A{row}'].font = Font(bold=True)
            worksheet[f'D{row}'] = absences_info['jours_deduits']
            worksheet[f'E{row}'] = f"{absences_info['deduction_montant']:.2f} DH"
            worksheet[f'E{row}'].font = Font(bold=True, color='D32F2F')
            row += 2
        
        # Tableau principal des gains et retenues
        worksheet[f'A{row}'] = "ÉLÉMENTS DE PAIE"
        worksheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        # Headers tableau principal
        headers = ['Désignation', 'Base', 'Taux', 'Montant (DH)']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col, value=header)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
        row += 1
        
        # Section GAINS
        gains_data = [
            ('Salaire de base', f'{bulletin.heures_travaillees}h', '', bulletin.salaire_base),
        ]
        
        if bulletin.total_primes > 0:
            gains_data.append(('Primes diverses', '—', '—', bulletin.total_primes))
        
        if bulletin.heures_supplementaires > 0:
            gains_data.append(('Heures supplémentaires', '—', '—', bulletin.heures_supplementaires))
        
        # Déduction absences si présente
        if absences_info['deduction_montant'] > 0:
            gains_data.append((
                'Déduction absences', 
                f"{absences_info['jours_deduits']}j", 
                '—', 
                -absences_info['deduction_montant']
            ))
        
        gains_data.append(('TOTAL BRUT IMPOSABLE', '', '', bulletin.salaire_brut_imposable))
        
        # Remplir les gains
        for designation, base, taux, montant in gains_data:
            worksheet[f'A{row}'] = designation
            worksheet[f'B{row}'] = base
            worksheet[f'C{row}'] = taux
            cell_montant = worksheet[f'D{row}']
            cell_montant.value = float(montant)
            cell_montant.number_format = self.styles['currency']
            
            # Style spécial pour déduction
            if montant < 0:
                cell_montant.font = Font(color='D32F2F')
            
            # Style pour total
            if 'TOTAL' in designation:
                for col in range(1, 5):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
            
            row += 1
        
        row += 1  # Espace
        
        # Section RETENUES
        worksheet[f'A{row}'] = "RETENUES"
        worksheet[f'A{row}'].font = Font(bold=True, color='D32F2F')
        row += 1
        
        retenues_data = [
            ('CNSS (4.48%)', f'{min(bulletin.salaire_brut_imposable, 6000):.2f}', '4.48%', bulletin.cotisation_cnss),
            ('AMO (2.26%)', f'{bulletin.salaire_brut_imposable:.2f}', '2.26%', bulletin.cotisation_amo),
            ('CIMR (6%)', f'{bulletin.salaire_brut_imposable:.2f}', '6%', bulletin.cotisation_cimr),
            ('Impôt sur le Revenu', 'Base imposable', 'Barème', bulletin.impot_revenu),
        ]
        
        if bulletin.total_retenues > 0:
            retenues_data.append(('Autres retenues', '—', '—', bulletin.total_retenues))
        
        if bulletin.total_avances > 0:
            retenues_data.append(('Avances sur salaire', '—', '—', bulletin.total_avances))
        
        total_retenues = (bulletin.cotisation_cnss + bulletin.cotisation_amo + 
                         bulletin.cotisation_cimr + bulletin.impot_revenu + 
                         bulletin.total_retenues + bulletin.total_avances)
        
        retenues_data.append(('TOTAL RETENUES', '', '', total_retenues))
        
        # Remplir les retenues
        for designation, base, taux, montant in retenues_data:
            worksheet[f'A{row}'] = designation
            worksheet[f'B{row}'] = base
            worksheet[f'C{row}'] = taux
            cell_montant = worksheet[f'D{row}']
            cell_montant.value = float(montant)
            cell_montant.number_format = self.styles['currency']
            
            if 'TOTAL' in designation:
                for col in range(1, 5):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
            
            row += 1
        
        row += 2  # Espace
        
        # NET À PAYER (encadré)
        worksheet[f'A{row}'] = "NET À PAYER"
        worksheet[f'A{row}'].font = Font(size=16, bold=True, color='2E7D32')
        
        cell_net = worksheet[f'D{row}']
        cell_net.value = float(bulletin.net_a_payer)
        cell_net.number_format = self.styles['currency']
        cell_net.font = Font(size=16, bold=True, color='2E7D32')
        
        # Bordure pour le net à payer
        for col in range(1, 5):
            cell = worksheet.cell(row=row, column=col)
            cell.border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='medium')
            )
            cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
        
        # Ajustement largeur colonnes
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 10
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 20
        
        # Sauvegarde
        workbook.save(output)
        output.seek(0)
        return output
    
    def export_bulletins_massif(self, bulletins_queryset, mois, annee):
        """
        Export massif de bulletins pour un mois donné
        """
        if not OPENPYXL_AVAILABLE:
            return None
            
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Bulletins_{mois:02d}_{annee}"
        
        # En-tête
        worksheet['A1'] = f"BULLETINS DE PAIE - {mois:02d}/{annee}"
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet.merge_cells('A1:M1')
        
        # Headers tableau
        headers = [
            'Matricule', 'Nom', 'Fonction', 'Salaire Base', 'Primes', 
            'Déduction Absences', 'Brut Imposable', 'CNSS', 'AMO', 'CIMR', 
            'IR', 'Total Retenues', 'Net à Payer'
        ]
        
        row = 3
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        row += 1
        
        # Données des bulletins
        total_net = 0
        total_brut = 0
        
        for bulletin in bulletins_queryset:
            # Calcul des absences pour chaque employé
            from .calculs import CalculateurPaie
            calculateur = CalculateurPaie()
            absences_info = calculateur.calculer_absences(
                bulletin.employe, bulletin.mois, bulletin.annee
            )
            
            data_row = [
                bulletin.employe.matricule,
                bulletin.employe.nom_complet(),
                bulletin.employe.fonction,
                float(bulletin.salaire_base),
                float(bulletin.total_primes),
                float(absences_info['deduction_montant']),
                float(bulletin.salaire_brut_imposable),
                float(bulletin.cotisation_cnss),
                float(bulletin.cotisation_amo),
                float(bulletin.cotisation_cimr),
                float(bulletin.impot_revenu),
                float(bulletin.cotisation_cnss + bulletin.cotisation_amo + 
                      bulletin.cotisation_cimr + bulletin.impot_revenu + 
                      bulletin.total_retenues + bulletin.total_avances),
                float(bulletin.net_a_payer)
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                
                # Format monétaire pour les colonnes financières
                if col >= 4:  # Colonnes financières
                    cell.number_format = '#,##0.00" DH"'
            
            total_net += bulletin.net_a_payer
            total_brut += bulletin.salaire_brut_imposable
            row += 1
        
        # Ligne de totaux
        row += 1
        worksheet[f'A{row}'] = "TOTAUX"
        worksheet[f'A{row}'].font = Font(bold=True)
        
        worksheet[f'G{row}'] = float(total_brut)
        worksheet[f'G{row}'].number_format = '#,##0.00" DH"'
        worksheet[f'G{row}'].font = Font(bold=True)
        
        worksheet[f'M{row}'] = float(total_net)
        worksheet[f'M{row}'].number_format = '#,##0.00" DH"'
        worksheet[f'M{row}'].font = Font(bold=True)
        
        # Ajustement largeurs
        for col in range(1, 14):
            worksheet.column_dimensions[get_column_letter(col)].width = 15
        
        workbook.save(output)
        output.seek(0)
        return output
    
    def export_format_cnss(self, bulletins_queryset, mois, annee):
        """
        Export au format CNSS pour déclarations officielles
        """
        if not OPENPYXL_AVAILABLE:
            return None
            
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"CNSS_{mois:02d}_{annee}"
        
        # En-tête CNSS
        worksheet['A1'] = f"DÉCLARATION CNSS - {mois:02d}/{annee}"
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet.merge_cells('A1:H1')
        
        # Headers format CNSS
        headers_cnss = [
            'N° d\'affiliation CNSS', 'CIN', 'Nom et Prénom', 'Nombre de jours',
            'Salaire Brut', 'Part Salariale CNSS', 'Part Patronale CNSS', 'Total CNSS'
        ]
        
        row = 3
        for col, header in enumerate(headers_cnss, 1):
            cell = worksheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        
        row += 1
        
        # Données CNSS
        total_cnss_salarie = 0
        total_cnss_patronal = 0
        
        for bulletin in bulletins_queryset:
            # Simulation numéro affiliation (à adapter selon vos données)
            num_affiliation = f"2020{bulletin.employe.matricule[-3:]}"
            
            # Part patronale CNSS (exemple: 20.48%)
            part_patronale = float(min(bulletin.salaire_brut_imposable, 6000)) * 0.2048
            
            data_cnss = [
                num_affiliation,
                bulletin.employe.cin,
                bulletin.employe.nom_complet(),
                26,  # Jours travaillés standard
                float(bulletin.salaire_brut_imposable),
                float(bulletin.cotisation_cnss),
                part_patronale,
                float(bulletin.cotisation_cnss) + part_patronale
            ]
            
            for col, value in enumerate(data_cnss, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                if col >= 5:  # Colonnes financières
                    cell.number_format = '#,##0.00" DH"'
            
            total_cnss_salarie += bulletin.cotisation_cnss
            total_cnss_patronal += part_patronale
            row += 1
        
        # Totaux CNSS
        row += 1
        worksheet[f'A{row}'] = "TOTAUX"
        worksheet[f'A{row}'].font = Font(bold=True)
        
        worksheet[f'F{row}'] = float(total_cnss_salarie)
        worksheet[f'F{row}'].number_format = '#,##0.00" DH"'
        worksheet[f'F{row}'].font = Font(bold=True)
        
        worksheet[f'G{row}'] = float(total_cnss_patronal)
        worksheet[f'G{row}'].number_format = '#,##0.00" DH"'
        worksheet[f'G{row}'].font = Font(bold=True)
        
        worksheet[f'H{row}'] = float(total_cnss_salarie + total_cnss_patronal)
        worksheet[f'H{row}'].number_format = '#,##0.00" DH"'
        worksheet[f'H{row}'].font = Font(bold=True)
        
        workbook.save(output)
        output.seek(0)
        return output
    
    def reponse_http_excel(self, excel_file, filename):
        """
        Retourne une réponse HTTP avec le fichier Excel
        """
        if not excel_file:
            return None
        
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response